'''Author       :       Annapoornima Koppad
Date            :       25 August 2016
Tite            :       Convert years of epxerience datetime to months
Purpose         :       This program reads excel sheet and converts years of ex$
'''

#import the necessary module to export excel sheet
import os, openpyxl
from openpyxl.cell import get_column_letter

path = os.getcwd()
excelfile='StudentsdataAugDec2016course29studentsData.xlsx'
file=path+'/'+excelfile

#open the workbook, its located in the same folder as this file
wb=openpyxl.load_workbook(file)

#the sheets in the excel sheet
sheets=wb.get_sheet_names()

#get to the first sheet where we want to read our values
sheet=wb.get_sheet_by_name(sheets[0])
sheet1=wb.get_sheet_by_name(sheets[1])

print sheet1.Name, "title"

sheet1.title="name"

print sheet1.title, "new name"
#get names and experince from the respective  columns
exp_column=get_column_letter(sheet.max_column)
student_name=get_column_letter(3)

#for name and experience column, get the row values for name and experience
experience_data={}

for i in range(1,sheet.max_row+1):
	namestr=student_name+str(i)
	expstr=exp_column+str(i)
	name, exp = sheet[namestr].value, sheet[expstr].value
	if name == None or name == '-Nil-' == name is '':
		continue
	else:
		if exp == None or exp == '-NIL-' or exp == '':
			continue
		else:
			#converting the experience data into months
			exp_d=exp.split()
			if len(exp_d)!=1:
				v=int(exp_d[0])*12+int(exp_d[2])
				print v
				experience_data[name]=v

for i in experience_data:
	print i, experience_data[i]


wb.save(file)
